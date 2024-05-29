import openpyxl
import itertools


def generate_permulations(text):
    result = []
    for i in itertools.permutations(text,len(text)):
        result.append(i)
    return result




def get_letter_probability(search_value):
    file_path = 'probability.xlsx'
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    # Перебираем все листы в книге
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        # Перебираем все строки и столбцы в листе
        for row in worksheet.iter_rows():
            for cell in row:
                # Если содержимое ячейки совпадает с искомым значением
                if cell.value == search_value:
                    # Получаем координаты текущей ячейки
                    row_idx = cell.row
                    col_idx = cell.column
                    # Получаем ячейку, смещенную на один столбец вправо
                    shifted_cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                    return shifted_cell.value
    # Если значение не найдено, возвращаем None
    return None
def get_transition_probability(header_value, index_value):
    file_path = 'transition_matrix.xlsx'
    # Открываем Excel файл
    workbook = openpyxl.load_workbook(file_path)
    # Выбираем активный лист
    worksheet = workbook.active

    # Ищем индекс столбца по значению в первой строке
    col_idx = None
    for cell in worksheet[1]:
        if cell.value == header_value:
            col_idx = cell.column
            print(col_idx)
            break

    # Ищем индекс строки по значению в первом столбце
    row_idx = None
    for cell in worksheet['A']:
        if cell.value == index_value:
            row_idx = cell.row
            print(row_idx)
            break

    # Если найдены оба индекса, возвращаем значение в ячейке на их пересечении
    if col_idx is not None and row_idx is not None:
        intersection_cell = worksheet.cell(row=row_idx, column=col_idx)
        return intersection_cell.value

    # Если значение не найдено, возвращаем None
    return None

# def markov_Chain(permutations):
#     for permutation in permutations:
#
#         markov = []
#         markov_p = []
#         for symbol in permutation:
#             char = symbol
#             char_p = get_letter_probability(char) #получаем вероятность буквы
#             markov.append(char)
#             markov_p.append(char_p)
#         print(markov)
#         print(markov_p)

def calculate_markov_chain_probability(permutations, text):

    # Функция для расчета вероятности слова
    def word_probability(word):
        if len(word) == 0:
            return 0

        # Начальная вероятность - вероятность первой буквы
        prob = float(get_letter_probability(word[0]))

        # Учитываем переходные вероятности между всеми парами букв
        for i in range(1, len(word)):
            transition_prob = get_transition_probability(word[i-1], word[i])
            prob *= float(transition_prob)
        return prob

    # Результаты для всех перестановок
    probabilities = []
    permutations_r = []

    for permutation in permutations:
        prob = word_probability(permutation)

        # probabilities.append(permutation)
        probabilities.append(prob)

    # return probabilities
    max_probability = max(probabilities)
    max_index = probabilities.index(max_probability)
    max_permutation = permutations[max_index]
    # file = open('result.txt', 'w')
    # file.write(str(max_probability))
    # file.write(str(max_permutation))
    # print(probabilities)

    if text == ''.join(max_permutation):
        print("Слово принадлежит русскому языку")
        return max_probability, max_permutation
    else:
        print("Скорее всего это не русское слово")
        return max_probability, max_permutation




text = input()
permutations = generate_permulations(text)
print(permutations)
result = calculate_markov_chain_probability(permutations, text)
print(result)
